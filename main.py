"""
Author: Samatha Ereshi Akkamahadevi
Email: samatha94@ksu.edu
Date: 08/27/2024
Project: Stage4: Statistical Analysis and Verification of Neuron Activations
File name: main.py
Description:
"""

# pip install pandas openpyxl
# pip install scipy

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from scipy.stats import mannwhitneyu

# Define the paths
# indices_file_path = 'C:\\Users\\Dell-PC\\OneDrive - Kansas State University\\Desktop\\DataSemantics prep\\ExAI\\outputs\\config_files\\indices.txt'
# evaluation_folder_path = 'C:\\Users\\Dell-PC\\OneDrive - Kansas State University\\Desktop\\DataSemantics prep\\ExAI\\outputs\\evaluation'
# config_files_path = 'C:\\Users\\Dell-PC\\OneDrive - Kansas State University\\Desktop\\DataSemantics prep\\ExAI\\outputs\\config_files'
# output_eval_path = 'C:\\Users\\Dell-PC\\OneDrive - Kansas State University\\Desktop\\DataSemantics prep\\ExAI\\outputs\\evaluation_combined.xlsx'
# verification_folder_path = 'C:\\Users\\Dell-PC\\OneDrive - Kansas State University\\Desktop\\DataSemantics prep\\ExAI\\outputs\\verification'
# output_veri_path = 'C:\\Users\\Dell-PC\\OneDrive - Kansas State University\\Desktop\\DataSemantics prep\\ExAI\\outputs\\verification_combined.xlsx'

indices_file_path = '/homes/samatha94/ExAI_inputs_and_outputs/Stage3_Results/indices.txt'
evaluation_folder_path = '/homes/samatha94/ExAI_inputs_and_outputs/Stage3_Results/evaluation'
config_files_path = '/homes/samatha94/ExAI_inputs_and_outputs/Stage1_Results/config_files'
output_eval_path = '/homes/samatha94/ExAI_inputs_and_outputs/Stage4_Results/evaluation_combined.xlsx'
verification_folder_path = '/homes/samatha94/ExAI_inputs_and_outputs/Stage3_Results/verification'
output_veri_path = '/homes/samatha94/ExAI_inputs_and_outputs/Stage4_Results/verification_combined.xlsx'


# Function to read coverage score from text file
def read_coverage_score(neuron_id, solution_number):
    filename = os.path.join(config_files_path, f"neuron_{neuron_id}_results_ecii_V2.txt")
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        solution_key = f'solution {solution_number}:'
        # Find the index of the line that contains the correct solution
        for i, line in enumerate(lines):
            if solution_key in line:
                # Ensure that the next line exists and contains 'coverage_score:'
                if i + 1 < len(lines) and 'coverage_score:' in lines[i + 1]:
                    coverage_line = lines[i + 1]
                    parts = coverage_line.split('coverage_score:')
                    if len(parts) > 1:
                        score_str = parts[1].split()[0]
                        return float(score_str)
                break  # Stop processing once the correct solution is processed

    except FileNotFoundError:
        print(f"File not found: {filename}")
    except Exception as e:
        print(f"An error occurred while reading from {filename}: {e}")

    return 0.0  # Default to 0 if no score found or error occurs



    
# Read neuron IDs from indices.txt
with open(indices_file_path, 'r') as file:
    neuron_ids = file.read().splitlines()

#----------------------------------------------------------------------------------------------------------------------------
#evaluation:
#----------------------------------------------------------------------------------------------------------------------------
# Initialize Excel writer with the output Excel file
with pd.ExcelWriter(output_eval_path, engine='openpyxl') as writer:
    # Process each solution type
    for solution_number in range(1, 4):
        combined_df = pd.DataFrame()
        # Process each neuron ID
        for neuron_id in neuron_ids:
            # Find matching CSV files for the current solution
            pattern = f"neuron{neuron_id}_solution{solution_number}_evaluation_set.csv"
            for filename in os.listdir(evaluation_folder_path):
                if pattern in filename:
                    # Read the CSV file
                    file_path = os.path.join(evaluation_folder_path, filename)
                    temp_df = pd.read_csv(file_path)
                    # Insert the filename as the first column
                    temp_df.insert(0, 'source name', filename)
                    # Append to the combined DataFrame
                    combined_df = pd.concat([combined_df, temp_df], ignore_index=True)
        
        # Write the combined DataFrame to the corresponding sheet in the Excel workbook
        combined_df.to_excel(writer, sheet_name=f'solution{solution_number}_all_activations', index=False)
        
        # Create summary DataFrame
        summary_df = pd.DataFrame(columns=['neuron_id', 'ecii concepts', 'coverage_score', 'target_activation', 'non_target_activation', 'target>20%', 'non target>20%', 'target>40%', 'non target>40%', 'target>60%', 'non target>60%'])
        
        for neuron_id in neuron_ids:
            # Filter data for target and non-target
            neuron_filter = combined_df['source name'].str.contains(f'neuron{neuron_id}')
            target_data = combined_df[neuron_filter][neuron_id].astype(float)
            non_target_data = combined_df[~neuron_filter][neuron_id].astype(float)

            # Basic activation calculations
            target_activation = (target_data > 0).sum() / len(target_data) * 100
            non_target_activation = (non_target_data > 0).sum() / len(non_target_data) * 100 if len(non_target_data) > 0 else 0
            coverage_score = read_coverage_score(neuron_id, solution_number)
            # Calculations for >20% of max
            max_target = target_data.max()
            max_non_target = non_target_data.max()
            target_above_20 = (target_data > max_target * 0.2).sum() / len(target_data) * 100 if len(target_data) > 0 else 0
            non_target_above_20 = (non_target_data > max_non_target * 0.2).sum() / len(non_target_data) * 100 if len(non_target_data) > 0 else 0
            target_above_40 = (target_data > max_target * 0.4).sum() / len(target_data) * 100 if len(target_data) > 0 else 0	
            non_target_above_40 = (non_target_data > max_non_target * 0.4).sum() / len(non_target_data) * 100 if len(non_target_data) > 0 else 0
            target_above_60 = (target_data > max_target * 0.6).sum() / len(target_data) * 100 if len(target_data) > 0 else 0	
            non_target_above_60 = (non_target_data > max_non_target * 0.6).sum() / len(non_target_data) * 100 if len(non_target_data) > 0 else 0
            
            # Extract the ecii concept/class name
            ecii_concept = combined_df[neuron_filter]['Class_names'].iloc[0] if not combined_df[neuron_filter].empty else 'Unknown'
            
            # Append to the summary DataFrame
            summary_df.loc[len(summary_df)] = [neuron_id, ecii_concept,coverage_score, target_activation, non_target_activation, target_above_20, non_target_above_20, target_above_40, non_target_above_40,target_above_60, non_target_above_60]
        
        # Write the summary DataFrame to the workbook
        summary_df.to_excel(writer, sheet_name=f'summary_{solution_number}', index=False)

        # Filter rows where target_activation > 80% and write to a new "final_summary" sheet
        final_summary_df = summary_df[summary_df['target_activation'] > 80]
        final_summary_df.to_excel(writer, sheet_name=f'final_summary_{solution_number}', index=False)

# Load workbook and apply formatting
wb = load_workbook(output_eval_path)
for solution_number in range(1, 4):
    ws = wb[f'summary_{solution_number}']
    # Apply bold formatting if target_activation > 80%
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        if row[3].value > 80:  # Assuming target_activation is in the third column
            for cell in row:
                cell.font = Font(bold=True)
wb.save(output_eval_path)

#----------------------------------------------------------------------------------------------------------------------------
#verification:
#----------------------------------------------------------------------------------------------------------------------------
# Read neuron IDs from indices.txt
with open(indices_file_path, 'r') as file:
    neuron_ids = file.read().splitlines()

# Load the final_summary data for reference
final_neuron_ids = []
for solution_number in range(1, 4):
    final_summary_df = pd.read_excel(output_eval_path, sheet_name=f'final_summary_{solution_number}')
    final_neuron_ids.append(final_summary_df['neuron_id'].unique())

# Process each solution type and create the activation and summary sheets sequentially
with pd.ExcelWriter(output_veri_path, engine='openpyxl') as writer:
    for solution_number in range(1, 4):
        combined_df = pd.DataFrame()
        # Process each neuron ID
        for neuron_id in neuron_ids:
            # Find matching CSV files for the current solution
            pattern = f"neuron{neuron_id}_solution{solution_number}_verification_set.csv"
            for filename in os.listdir(verification_folder_path):
                if pattern in filename:
                    # Read the CSV file
                    file_path = os.path.join(verification_folder_path, filename)
                    temp_df = pd.read_csv(file_path)
                    # Insert the filename as the first column
                    temp_df.insert(0, 'source name', filename)
                    # Append to the combined DataFrame
                    combined_df = pd.concat([combined_df, temp_df], ignore_index=True)

        # Write the combined DataFrame to the corresponding sheet in the Excel workbook
        combined_df.to_excel(writer, sheet_name=f'solution{solution_number}_all_activations', index=False)

        # Prepare the summary DataFrame
        summary_df = pd.DataFrame(columns=['neuron_id', 'ecii concepts', 'target_activation', 'non_target_activations','target_median', 'non_target_median', 'target_mean', 'non_target_mean','z_score', 'p_value'])

        for neuron_id in final_neuron_ids[solution_number - 1]:
            neuron_filter = combined_df['source name'].str.contains(f'neuron{neuron_id}')
            target_data = combined_df[neuron_filter][str(neuron_id)].astype(float)
            non_target_data = combined_df[~neuron_filter][str(neuron_id)].astype(float)

            # Calculate activations and statistical measures
            target_activation = (target_data > 0).sum() / len(target_data) * 100 if len(target_data) > 0 else 0
            non_target_activation = (non_target_data > 0).sum() / len(non_target_data) * 100 if len(non_target_data) > 0 else 0
            target_median = target_data.median()
            non_target_median = non_target_data.median()
            target_mean = target_data.mean()
            non_target_mean = non_target_data.mean()

            # Mann-Whitney U test
            u_stat, p_val = mannwhitneyu(target_data, non_target_data, alternative='two-sided')
            z_score = (u_stat - (len(target_data) * len(non_target_data)/2)) / (len(target_data) * len(non_target_data) * (len(target_data) + len(non_target_data) + 1) / 12)**0.5


            # Extract the ecii concept/class name
            ecii_concept = combined_df[neuron_filter]['Class_names'].iloc[0] if not combined_df[neuron_filter].empty else 'Unknown'

            # Append to the summary DataFrame
            summary_df.loc[len(summary_df)] = {
                'neuron_id': neuron_id,
                'ecii concepts': ecii_concept,
                'target_activation': target_activation,
                'non_target_activations': non_target_activation,
                'target_median': target_median,
                'non_target_median': non_target_median,
                'target_mean': target_mean,
                'non_target_mean': non_target_mean,
                'z_score': z_score, 'p_value': p_val
            }

        # Write the summary DataFrame to the new summary sheet
        summary_df.to_excel(writer, sheet_name=f'summary_{solution_number}', index=False)


# Apply formatting based on p-value
wb = load_workbook(output_veri_path)
for solution_number in range(1, 4):
    ws = wb[f'summary_{solution_number}']
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        if row[-1].value < 0.05:  # Assuming p_value is the last column
            for cell in row:
                cell.font = Font(bold=True)
wb.save(output_veri_path)